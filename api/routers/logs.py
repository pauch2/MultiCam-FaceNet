"""
logs.py — Detection & unknown-face logs with filtering and Excel export.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import desc
from sqlalchemy.orm import Session

from api.auth import require_role, get_current_user
from database.models import Camera, Detection, UnknownFace
from database.session import get_db

router = APIRouter(prefix="/logs", tags=["Logs"])


# ── helpers ───────────────────────────────────────────────────────────────────

def _cam_name_map(db: Session) -> dict[int, str]:
    return {c.id: c.name for c in db.query(Camera).all()}


def _apply_det_filters(q, camera_db_id, name, date_from, date_to):
    """Apply optional filters to a Detection query."""
    if camera_db_id is not None:
        q = q.filter(Detection.camera_db_id == camera_db_id)
    if name:
        q = q.filter(Detection.name.ilike(f"%{name}%"))
    if date_from:
        try:
            q = q.filter(Detection.timestamp >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(Detection.timestamp <= datetime.fromisoformat(date_to + "T23:59:59"))
        except ValueError:
            pass
    return q


def _apply_unk_filters(q, camera_db_id, date_from, date_to):
    if camera_db_id is not None:
        q = q.filter(UnknownFace.camera_db_id == camera_db_id)
    if date_from:
        try:
            q = q.filter(UnknownFace.timestamp >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(UnknownFace.timestamp <= datetime.fromisoformat(date_to + "T23:59:59"))
        except ValueError:
            pass
    return q


def _fmt(ts) -> str | None:
    return ts.strftime("%Y-%m-%d %H:%M:%S") if ts else None


# ── Detection log ─────────────────────────────────────────────────────────────

def _det_rows(rows, cam_names):
    return [dict(
        id           = d.id,
        name         = d.name or "Unknown",
        camera_id    = d.camera_id,
        camera_db_id = d.camera_db_id,
        camera_name  = cam_names.get(d.camera_db_id, d.camera_id or "-"),
        timestamp    = _fmt(d.timestamp),
        confidence   = d.confidence,
        image_path   = d.image_path,
    ) for d in rows]


@router.get("/detections")
def get_detections(
    limit:        int           = Query(100, le=5000),
    camera_db_id: Optional[int] = Query(None),
    name:         Optional[str] = Query(None),
    date_from:    Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_to:      Optional[str] = Query(None, description="YYYY-MM-DD"),
    db:           Session       = Depends(get_db),
    current_user               = Depends(require_role(["admin", "moderator"])),
):
    q = db.query(Detection).order_by(desc(Detection.timestamp))
    q = _apply_det_filters(q, camera_db_id, name, date_from, date_to)
    rows = q.limit(limit).all()
    return {"status": "success", "data": _det_rows(rows, _cam_name_map(db))}


@router.get("/my")
def get_my_detections(
    limit:        int           = Query(200, le=2000),
    camera_db_id: Optional[int] = Query(None),
    date_from:    Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_to:      Optional[str] = Query(None, description="YYYY-MM-DD"),
    db:           Session       = Depends(get_db),
    current_user               = Depends(get_current_user),
):
    """Returns detections only for the logged-in user (all roles can access).
    Filters by user_id (integer FK) — immune to display_name collisions.
    """
    from database.models import User as UserModel
    user_row = db.query(UserModel).filter(UserModel.username == current_user.username).first()
    if not user_row:
        return {"status": "success", "data": []}
    q = db.query(Detection).order_by(desc(Detection.timestamp))
    q = q.filter(Detection.user_id == user_row.id)
    q = _apply_det_filters(q, camera_db_id, None, date_from, date_to)
    rows = q.limit(limit).all()
    return {"status": "success", "data": _det_rows(rows, _cam_name_map(db))}


# ── Unknown faces ─────────────────────────────────────────────────────────────

@router.get("/unknown")
def get_unknown_faces(
    limit:        int           = Query(100, le=5000),
    camera_db_id: Optional[int] = Query(None),
    date_from:    Optional[str] = Query(None),
    date_to:      Optional[str] = Query(None),
    db:           Session       = Depends(get_db),
    current_user               = Depends(require_role(["admin", "moderator"])),
):
    q = db.query(UnknownFace).order_by(desc(UnknownFace.timestamp))
    q = _apply_unk_filters(q, camera_db_id, date_from, date_to)
    rows = q.limit(limit).all()
    cam_names = _cam_name_map(db)

    data = [dict(
        id           = d.id,
        camera_id    = d.camera_id,
        camera_db_id = d.camera_db_id,
        camera_name  = cam_names.get(d.camera_db_id, d.camera_id or "-"),
        timestamp    = _fmt(d.timestamp),
        image_path   = d.image_path,
    ) for d in rows]

    return {"status": "success", "data": data}


# ── Excel export ──────────────────────────────────────────────────────────────

@router.get("/export/xlsx")
def export_xlsx(
    camera_db_id: Optional[int] = Query(None),
    name:         Optional[str] = Query(None),
    date_from:    Optional[str] = Query(None),
    date_to:      Optional[str] = Query(None),
    db:           Session       = Depends(get_db),
    current_user               = Depends(require_role(["admin", "moderator"])),
):
    """Export filtered detections + unknown faces to a formatted .xlsx file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Fetch data ────────────────────────────────────────────────────────────
    cam_names = _cam_name_map(db)

    q_det = db.query(Detection).order_by(desc(Detection.timestamp))
    q_det = _apply_det_filters(q_det, camera_db_id, name, date_from, date_to)
    detections = q_det.limit(5000).all()

    q_unk = db.query(UnknownFace).order_by(desc(UnknownFace.timestamp))
    q_unk = _apply_unk_filters(q_unk, camera_db_id, date_from, date_to)
    unknowns = q_unk.limit(5000).all()

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Style helpers
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ALT_FILL    = PatternFill("solid", fgColor="F2F6FC")
    BORDER_SIDE = Side(style="thin", color="D0D0D0")
    THIN_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                         top=BORDER_SIDE, bottom=BORDER_SIDE)
    CENTER      = Alignment(horizontal="center", vertical="center")
    WRAP        = Alignment(wrap_text=True, vertical="center")

    def _style_sheet(ws, headers: list[str], col_widths: list[int]):
        ws.freeze_panes = "A2"
        for ci, (hdr, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(1, ci, hdr)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = CENTER
            cell.border    = THIN_BORDER
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 22

    def _add_row(ws, row_i: int, values: list):
        fill = ALT_FILL if row_i % 2 == 0 else None
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row_i, ci, v)
            cell.border    = THIN_BORDER
            cell.alignment = WRAP
            if fill:
                cell.fill = fill

    # ── Sheet 1: Recognised detections ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Recognised"
    headers1 = ["ID", "Name", "Camera", "Timestamp", "Confidence"]
    widths1  = [8, 22, 20, 22, 14]
    _style_sheet(ws1, headers1, widths1)

    for i, d in enumerate(detections, 2):
        cam = cam_names.get(d.camera_db_id, d.camera_id or "-")
        conf = round(float(d.confidence), 4) if d.confidence is not None else None
        _add_row(ws1, i, [d.id, d.name or "Unknown", cam, _fmt(d.timestamp), conf])

    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}1"

    # ── Sheet 2: Unknown faces ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Unknown Faces")
    headers2 = ["ID", "Camera", "Timestamp"]
    widths2  = [8, 24, 22]
    _style_sheet(ws2, headers2, widths2)

    for i, d in enumerate(unknowns, 2):
        cam = cam_names.get(d.camera_db_id, d.camera_id or "-")
        _add_row(ws2, i, [d.id, cam, _fmt(d.timestamp)])

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}1"

    # ── Sheet 3: Filter summary ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Export Info")
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 32
    ws3.cell(1, 1, "Export generated").font = Font(bold=True)
    ws3.cell(1, 2, datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"))
    params = [
        ("Camera filter", cam_names.get(camera_db_id, str(camera_db_id)) if camera_db_id else "All"),
        ("Name filter",   name or "All"),
        ("Date from",     date_from or "—"),
        ("Date to",       date_to   or "—"),
        ("Recognised rows",  len(detections)),
        ("Unknown rows",     len(unknowns)),
    ]
    for ri, (k, v) in enumerate(params, 2):
        ws3.cell(ri, 1, k).font = Font(bold=True)
        ws3.cell(ri, 2, str(v))

    # ── Stream response ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = "detections"
    if date_from: fname += f"_{date_from}"
    if date_to:   fname += f"_to_{date_to}"
    fname += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )